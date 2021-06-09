from math import fabs
from os import name
import subprocess
import sys
import openpyxl

import datetime
import numpy as np
import copy

import pandas as pd
import PyQt5 
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QObject, Qt, pyqtSignal
from PyQt5.QtWidgets import *
import random
import urllib.request
import json
bingMapsKey = "z04OKoL4xG4wPX3Qht8u~kN613p_8YYryf8s86kHfLA~AoIt5JF8V791QKh5Ty85LCwq-qRoCHRSPsZE2OE43KdRA1mdBu8Me6Ufq58keovG"
depot  = 0 
N = 50
t_max = 3
generations = 30
numChrosome = 50
p_c = 0.85
p_m = 0.25
demand = []

vehi_cap = []
cap = 0
num_vehi = 0
distance_mt = []
error = ''
class chromosome():
    def __init__(self,nCusPoint,presentation):
        self.nCusPoint = nCusPoint
        self.presentation = []
        self.nowloading = []
        self.fitness = 0
        for i in range(len(presentation)):
            tmpR = []
            tmpLoading  = 0
            if len(presentation[i]) != 1:
                
                for j in range(len(presentation[i])):
                    tmpR.append(presentation[i][j])
                    tmpLoading += demand[presentation[i][j]]
                    if j != 0 and j != len(presentation[i]) - 1:
                        self.fitness += distance_mt[presentation[i][j-1]][presentation[i][j]]
                    if j == len(presentation[i]) - 1 :
                        self.fitness += distance_mt[presentation[i][j-1]][presentation[i][j]]
                        self.fitness += distance_mt[presentation[i][j]][0]
                    if j == 0:
                        self.fitness += distance_mt[0][presentation[i][j]]
                self.presentation.append(tmpR)
                self.nowloading.append(tmpLoading)
            else:
                tmpR.append(presentation[i][0])
                self.fitness += distance_mt[0][presentation[i][0]] + distance_mt[presentation[i][0]][0]
                tmpLoading += demand[presentation[i][0]]
                self.presentation.append(tmpR)
                self.nowloading.append(tmpLoading)
    def checkError(self):
        # print(demand)
        # print(self.presentation)
        for i in range(len(self.presentation)):
            tmpLoad = 0
            for j in range(len( self.presentation[i])):
                tmpLoad+= demand[self.presentation[i][j]]
            # print(str(tmpLoad) + '%' + str(cap))
            if tmpLoad > cap:
                return True
        return False
    def updateFitness(self):
        self.nowloading = []
        self.fitness = 0
        for i in range(len(self.presentation)):
            tmpLoading  = 0
            if len(self.presentation[i]) != 1:
                for j in range(len(self.presentation[i])):
                    tmpLoading += demand[self.presentation[i][j]]
                    if j != 0 and j != len(self.presentation[i]) - 1:
                        self.fitness += distance_mt[self.presentation[i][j-1]][self.presentation[i][j]]
                    if j == len(self.presentation[i]) - 1 :
                        self.fitness += distance_mt[self.presentation[i][j-1]][self.presentation[i][j]]
                        self.fitness += distance_mt[self.presentation[i][j]][0]
                    if j == 0:
                        self.fitness += distance_mt[0][self.presentation[i][j]]
                self.nowloading.append(tmpLoading)
            else:
                self.fitness += distance_mt[0][self.presentation[i][0]] + distance_mt[self.presentation[i][0]][0]
                tmpLoading += demand[self.presentation[i][0]]
                self.nowloading.append(tmpLoading)
        
class population():
    def __init__(self,nChromo,nCusPoint):
        self.nChromo = nChromo
        self.nCusPoint = nCusPoint
        self.pop = []
        self.bestFit = 1000000000000
    def addChromosome(self,chromosome):
        self.pop.append(chromosome)
    def totalFitness(self):
        total = 0
        for i in range(self.nChromo):
            total += self.pop[i]
        return total
    def bestFitness(self):
        
        for i in range(numChrosome):
            if self.pop[i].fitness < self.bestFit:
                self.bestFit = self.pop[i].fitness
                self.bestChro = self.pop[i]

    def crossover(self):
        self.filialPop = []
        for i in range(numChrosome):
            crossRate = random.random()
            if crossRate < p_c:
                #select father and mother
                father_index = 1
                bestFitInFather = 1000000
                for j in range(5):
                    tmpFatherIndex = random.randint(0,len(self.pop)-1)
                    if self.pop[tmpFatherIndex].fitness < bestFitInFather:
                        bestFitInFather = self.pop[tmpFatherIndex].fitness
                        father_index = tmpFatherIndex
                mother_index = 1
                bestFitInMother = 1000000
                for j in range(5):
                    tmpMotherIndex = random.randint(0,len(self.pop)-1)
                    if self.pop[mother_index].fitness < bestFitInMother:
                        bestFitInMother = self.pop[mother_index].fitness
                        mother_index = tmpMotherIndex
                #create binary system
                LengOfBinarySystem = num_vehi
                binarySystem  = []
                for j in range(LengOfBinarySystem):
                    tmpBinary = random.randint(0,1)
                    binarySystem.append(tmpBinary)
                #create offstring 1 and 2
                offstring1 = []
                for j in range(num_vehi):
                    tmpRoute = []
                    offstring1.append(tmpRoute)
                offstring2 = []
                for j in range(num_vehi):
                    tmpRoute = []
                    offstring2.append(tmpRoute)
                list0InBinarySys1  = []
                list0InBinarySys2  = []
                nowLoading1 = []
                for j in range(num_vehi):
                    nowLoading1.append(0)
                nowLoading2 = []
                for j in range(num_vehi):
                    nowLoading2.append(0)
                #    
                for j in range(LengOfBinarySystem):
                    #copy "1" route to offstring
                    if binarySystem[j] == 1 :
                        for k in range(len(self.pop[father_index].presentation[j])):
                            offstring1[j].append(self.pop[father_index].presentation[j][k])
                            nowLoading1[j] += demand[self.pop[father_index].presentation[j][k]]
                        for k in range(len(self.pop[mother_index].presentation[j])):
                            offstring2[j].append(self.pop[mother_index].presentation[j][k])
                            nowLoading2[j] += demand[self.pop[mother_index].presentation[j][k]]
                    
                    else:
                        for k in self.pop[father_index].presentation[j]:
                            list0InBinarySys1.append(k)
                        for k in self.pop[mother_index].presentation[j]:
                            list0InBinarySys2.append(k)
                # with "0" route form into new order according to another parent
                for j in range(len(self.pop[mother_index].presentation)):
                    checkOver = True
                    for k in self.pop[mother_index].presentation[j]:
                        if k in list0InBinarySys1:
                            for t in range(num_vehi):
                                checkOver = False
                                if demand[k] + nowLoading1[t] <= cap:
                                    offstring1[t].append(k)
                                    nowLoading1[t] += demand[k]
                                    checkOver = True
                                    break
                            if checkOver == False:
                                offstring1 = copy.deepcopy(self.pop[father_index].presentation)
                                break
                    if checkOver == False:
                        break
                for j in range(len(self.pop[father_index].presentation)):
                    checkOver = True
                    for k in self.pop[father_index].presentation[j]:
                        if k in list0InBinarySys2:
                            for t in range(num_vehi):
                                checkOver = False
                                if demand[k] + nowLoading2[t] <= cap:
                                    offstring2[t].append(k)
                                    nowLoading2[t] += demand[k]
                                    checkOver = True
                                    break
                            if checkOver == False:
                                offstring2 = copy.deepcopy(self.pop[mother_index].presentation)
                                break
                    if checkOver == False:
                        break
                
                filialChromo1 = chromosome(self.nCusPoint,offstring1)
                filialChromo2 = chromosome(self.nCusPoint,offstring2)
                # if filialChromo1.checkError() or filialChromo2.checkError():
                #     exit()
                self.filialPop.append(filialChromo1)
                self.filialPop.append(filialChromo2)
            
    
    def mutation(self):
   
        for i in range(numChrosome):
            mutationRate = random.random()
            if mutationRate < p_m:
                copy1 = copy.deepcopy(self.pop[i].presentation)
                # swap 2 place in chromosome if satifysing condition
                firstRouteToMutation = random.randint(0,num_vehi-1)# Route  choose to swap
                secondRouteToMutation = random.randint(0,num_vehi-1)# Route  choose to swap
                while len(self.pop[i].presentation[firstRouteToMutation]) == 0:
                    firstRouteToMutation = random.randint(0,num_vehi-1)
                firstPlaceToMutation = random.randint(0,len(self.pop[i].presentation[firstRouteToMutation])-1)# Point choose to swap
                while len(self.pop[i].presentation[secondRouteToMutation]) == 0:
                    secondRouteToMutation = random.randint(0,num_vehi-1)
                secondPlaceToMutation = random.randint(0,len(self.pop[i].presentation[secondRouteToMutation])-1)# Point choose to swap
                firstPoint = self.pop[i].presentation[firstRouteToMutation][firstPlaceToMutation]
                secondPoint  = self.pop[i].presentation[secondRouteToMutation][secondPlaceToMutation]
                while True:
                    if self.pop[i].nowloading[firstRouteToMutation] - demand[firstPoint] + demand[secondPoint] <= cap and self.pop[i].nowloading[secondRouteToMutation] - demand[secondPoint] + demand[firstPoint] <= cap:
                        break
                    else:
                        firstRouteToMutation = random.randint(0,num_vehi-1) # Route  choose to swap
                        while len(self.pop[i].presentation[firstRouteToMutation]) == 0:
                            firstRouteToMutation = random.randint(0,num_vehi-1)
                        secondRouteToMutation = random.randint(0,num_vehi-1) # Route  choose to swap
                        while len(self.pop[i].presentation[secondRouteToMutation]) == 0:
                            secondRouteToMutation = random.randint(0,num_vehi-1)
                        firstPlaceToMutation = random.randint(0,len(self.pop[i].presentation[firstRouteToMutation]) - 1) # Point choose to swap
                        secondPlaceToMutation = random.randint(0,len(self.pop[i].presentation[secondRouteToMutation])-1)# Point choose to swap
                        firstPoint = self.pop[i].presentation[firstRouteToMutation][firstPlaceToMutation]
                        secondPoint  = self.pop[i].presentation[secondRouteToMutation][secondPlaceToMutation]
                tmpSwap = self.pop[i].presentation[firstRouteToMutation][firstPlaceToMutation]
                self.pop[i].presentation[firstRouteToMutation][firstPlaceToMutation] = self.pop[i].presentation[secondRouteToMutation][secondPlaceToMutation]
                self.pop[i].presentation[secondRouteToMutation][secondPlaceToMutation] = tmpSwap
                self.pop[i].updateFitness() 
                # if self.pop[i].checkError():
                #     exit()
                    
    def survivorSelection(self):
        # Fitness based selection
        tmpParentPop = copy.deepcopy(self.pop)
        tmpFilialPop = copy.deepcopy(self.filialPop)
        tmpPop = tmpParentPop + tmpFilialPop
        emptyPop = []
        numberOfTotalPop = len(tmpFilialPop) + len(tmpParentPop)
        
        
        for i in range(len(tmpPop)):
           for j in range(i+1,len(tmpPop)):
               if tmpPop[i].fitness > tmpPop[j].fitness:
                   tmpFit = tmpPop[i]
                   tmpPop[i] = tmpPop[j]
                   tmpPop[j] = tmpFit
        limitIndex = 0
        for i in range(len(tmpPop)):
            if tmpPop[i].fitness > tmpPop[0].fitness:
                limitIndex = i
                break
        # Chromosome with fitness <= limit fitness will be choose
        for i in range(numChrosome-1,numChrosome-1+int(numChrosome/6)):
            chooseIndex = random.randint(limitIndex,len(tmpPop) - 1)
            tmpChromo = copy.deepcopy(tmpPop[chooseIndex])
            emptyPop.append(tmpChromo)
            chooseIndex -= 1
        for i in range(numChrosome - int(numChrosome/6)):
            tmpChromo = copy.deepcopy(tmpPop[i])
            emptyPop.append(tmpChromo)
        
        
        
        self.pop = copy.deepcopy(emptyPop)

def isnan(value):
    try:
        import math
        return math.isnan(float(value))
    except:
        return True
def create_data_model():
    """Stores the data for the problem."""
    x = pd.read_excel('data/data.xlsx',0,header=0, index_col= False)
    sheet_2 = pd.read_excel('data/data.xlsx',sheet_name = "Danh sách xe",header=None, index_col= False)
    data = {}

    title_sheet2 = sheet_2.columns.ravel()
    data['list_cap'] = []
    for t in title_sheet2:
        for k in sheet_2[t].tolist():
            if isnan(k) == False:
                data['list_cap'].append(k)
    data['list_cap'].reverse()
    y = x.columns.ravel()
    z = []

    data['wd_names'] = []
    for t in range(len(y)):
        if (y[t].find('Unnamed') == -1):
            if (t > 1):
                z.append(x[y[t]].tolist())
            
            data['wd_names'].append(y[t])
    count_col = 0
    for i in range(len(z)):
        count_col = 0
        while True:
            if not (isnan(z[i][count_col])):
                count_col+=1
            else:
                z[i].remove(z[i][count_col])
            if count_col == len(z[i]) :
                break

    
    for v in z:
        for k in range(len(v)):
            if isnan(v[k]):
                v[k] = 0
    z = np.array(z)
    z = z.T
    z = z.tolist()
    data['distance_matrix'] = z
    data['distance_matrix'] = ( [list( map(int,i) ) for i in data['distance_matrix']] )
    data['demands'] = []
    for p in range(len(data['distance_matrix'])):
        data['demands'].append(0)
    
    data['wd_names_now'] = []
    return data

data1 = create_data_model()
or_data = create_data_model()
rs = []
def modify_data(data,check_run):
    data['wd_names_now'] = []
    goc_appear = False
    data['ori_len'] = len(data['demands'])
    demand1 = copy.deepcopy(data['demands'])
    distance_matrix1 = copy.deepcopy(data['distance_matrix'])
    data['demands'] = [0]
    data['distance_matrix'] = []
    tmp1 =[0]
    data['distance_matrix'].append(tmp1)
    
    index_need_demand= []
    count = 0
    if demand1[0] == 0:
        goc_appear = False
    else:
        goc_appear = True
    if goc_appear == False:
        data['wd_names_now'].append(data['wd_names'][0])

        for p in range(len(demand1)):
            if demand1[p] != 0:
                count+=1

                data['demands'].append(demand1[p])
                data['wd_names_now'].append(data['wd_names'][p])

                data['distance_matrix'][0].append(distance_matrix1[0][p])
                tmp = []
                tmp.append(distance_matrix1[p][0])
                data['distance_matrix'].append(tmp)
                tmp_dic = {'pre' : p, 'now' : count}
                index_need_demand.append(tmp_dic)
        for p in range(1,len(data['distance_matrix'])):
            for t in range(1,len(data['distance_matrix'][0])):
                if t == p:
                    data['distance_matrix'][p].append(0)
                else:
                    for v in index_need_demand:
                        if v['now'] == p:
                            row = v['pre']
                        if v['now'] == t:
                            col = v['pre']
                    
                    
                    data['distance_matrix'][p].append(distance_matrix1[row][col])
    else:
        data['wd_names_now'].append(data['wd_names'][0])

        for p in range(1,len(demand1)):
            if demand1[p] != 0:
                count+=1
        
                data['demands'].append(demand1[p])
                data['wd_names_now'].append(data['wd_names'][p])

                data['distance_matrix'][0].append(distance_matrix1[0][p])
                tmp = []
                tmp.append(distance_matrix1[p][0])
                data['distance_matrix'].append(tmp)
                tmp_dic = {'pre' : p, 'now' : count}
                index_need_demand.append(tmp_dic)
        for p in range(1,len(data['distance_matrix'])):
            for t in range(1,len(data['distance_matrix'][0])):
                if t == p:
                    data['distance_matrix'][p].append(0)
                else:
                    for v in index_need_demand:
                        if v['now'] == p:
                            row = v['pre']
                        if v['now'] == t:
                            col = v['pre']
                    data['distance_matrix'][p].append(distance_matrix1[row][col])
    data['vehicle_capacities'] = []
    data['num_vehicles'] = 0
    data['depot'] = 0
    total_demand = 0
    if check_run == True:
        data['du_thua'] = copy.deepcopy(data['demands'])
        tmp_divine = 0
        for i in range(len(data['demands'])):
            if data['demands'][i] > data['list_cap'][0]:
                if (data['demands'][i] % data['list_cap'][0] != 0):
                    tmp_divine = data['demands'][i]//data['list_cap'][0]
                    data['demands'][i] = data['demands'][i] - data['list_cap'][0]*tmp_divine
                    data['du_thua'][i] = tmp_divine

                else:
                    tmp_divine = (data['demands'][i]//data['list_cap'][0]) - 1
                    data['demands'][i] = data['list_cap'][0]
                    data['du_thua'][i] = tmp_divine

            else:
                data['du_thua'][i] = 0

    copy_demand = copy.deepcopy(data['demands'])

    for i in data['demands']:
        total_demand += i
    total_cap = 0
    check = True
    posi_cap = 0

    copy_demand.sort(reverse=True)
    nowload = []
    for i in range(len(copy_demand)):
        checkMaintain = False
        for j in range(len(nowload)):
            if nowload[j] + copy_demand[i] <= data['list_cap'][0]:
                nowload[j] += copy_demand[i]
                checkMaintain = True
                break
        if checkMaintain == False :
            nowload.append(copy_demand[i])
    for i in range(len(nowload)):
        data['vehicle_capacities'].append(data['list_cap'][0])
    
    global demand
    global vehi_cap
    global cap
    global distance_mt
    global num_vehi
    demand = data['demands']
    vehi_cap = data['vehicle_capacities']
    cap = vehi_cap[0]
    num_vehi = len(vehi_cap)
    distance_mt = data['distance_matrix']


    return data['du_thua']
# [0, 1, 1, 3, 3, 2, 4, 13, 7, 1, 2, 1, 2, 4, 1, 4, 4, 4]
def print_solution(data, presentation):
    """Prints solution on console."""
    wb_obj = openpyxl.load_workbook("result/result.xlsx")
    wb_obj.save("result/result.xlsx")
    wb_obj = openpyxl.load_workbook("result/result.xlsx")
    sheet_obj = wb_obj.active
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row
    result = []
    count_vehi = 0
    temp_vehi = copy.deepcopy(data1['vehicle_capacities'])
    tmpstr = 'Danh sách địa điểm:'
    result.append(tmpstr)
    for tmpIndex in range(len(data['wd_names_now'])):
        tmpstr = '{0}: {1}'.format(tmpIndex, data['wd_names_now'][tmpIndex])
        # name_cell=sheet_obj.cell(row=tmpIndex+max_row+1,column=1)
        # name_cell.value = tmpIndex+1
        result.append(tmpstr)
    total_distance = 0
    total_load = 0
    vehicle_id = 0
    loaed = []
    new_cap = []
    for i in range(len(presentation)): 
        loaed.append(0)
        for j in range(len(presentation[i])):
            loaed[i] +=demand[presentation[i][j]]
        for j in range(len(data['list_cap'])):
            if j != len(data['list_cap']) - 1:
                if loaed[i] <= data['list_cap'][j] and loaed[i] > data['list_cap'][j+1]:
                    new_cap.append(data['list_cap'][j])
                    break
            else:
                new_cap.append(data['list_cap'][-1])
                break
    stt = 1
    index_row = max_row+1
    tmpstr = '            '
    result.append(tmpstr)
    for vehicle_id in range(len(presentation)):
        nowload = 0
        route_distance = 0
        tmpstr = 'Lộ trình {0} - xe: {1} KGS'.format(vehicle_id,new_cap[vehicle_id])
        result.append(tmpstr)
        tmpstr = ''
        tmpstr += 'Từ {0} đến {1}:'.format(0,presentation[vehicle_id][0])
        tmpstr += ' Khối lượng: {0} KGS | Quãng đường: {1} km'.format(demand[presentation[vehicle_id][0]],distance_mt[0][presentation[vehicle_id][0]])
        result.append(tmpstr)
        j =0
        for j in range(len(presentation[vehicle_id])):
            tmpstr = ''
            nowload += demand[presentation[vehicle_id][j]]
            if (j != 0):
                tmpstr +='Từ {0} đến {1}:'.format(presentation[vehicle_id][j-1],presentation[vehicle_id][j])
                tmpstr += ' Khối lượng: {0} KGS | Quãng đường: {1} km'.format(demand[presentation[vehicle_id][j]],distance_mt[presentation[vehicle_id][j-1]][presentation[vehicle_id][j]])
                result.append(tmpstr)
                # tmpstr += '{0} Vận chuyển({1} KGS) ->'.format(presentation[vehicle_id][j],nowload)
            # write result to result.xlsx
            stt_cell=sheet_obj.cell(row=index_row,column=1)
            stt_cell.value = stt
            month_cell = sheet_obj.cell(row=index_row,column=2)
            x = datetime.datetime.now()
            month_cell.value = x.strftime("%b") + "-" + x.strftime("%y")
            date_cell = sheet_obj.cell(row=index_row,column=3)
            date_cell.value = x.strftime("%d")
            pickUp_cell = sheet_obj.cell(row=index_row,column=4)
            pickUp_cell.value = data['wd_names_now'][0]
            deli_cell = sheet_obj.cell(row=index_row,column=5)
            deli_cell.value = data['wd_names_now'][presentation[vehicle_id][j]]
            weight_cell = sheet_obj.cell(row=index_row,column=9)
            weight_cell.value = demand[presentation[vehicle_id][j]]
            route_cell = sheet_obj.cell(row = index_row, column = 6 )
            route_cell.value = "Route" + str(vehicle_id+1)
            orderRoute_cell = sheet_obj.cell(row = index_row, column = 7)
            orderRoute_cell.value = j+1
            capactiy_cell = sheet_obj.cell(row = index_row, column = 10)
            capactiy_cell.value  = new_cap[vehicle_id]
            
            #

            if j == 0:
                route_distance += distance_mt[0][presentation[vehicle_id][j]]
            else:
                route_distance += distance_mt[presentation[vehicle_id][j-1]][presentation[vehicle_id][j]]
            if j == len(presentation[vehicle_id])-1:
                distance_cell = sheet_obj.cell(row = index_row,column = 8)
                distance_cell.value = route_distance + distance_mt[presentation[vehicle_id][j]][0]
            index_row += 1
            stt += 1
        # print(str(vehicle_id) + '-' + str(j))
        route_distance += distance_mt[presentation[vehicle_id][j]][0]
        total_distance += route_distance
        total_load += nowload
        tmpstr = ''
        tmpstr +='Từ {0} đến {1}:'.format(presentation[vehicle_id][j],0)
        tmpstr += ' Khối lượng: {0} KGS | Quãng đường: {1} km'.format(0,distance_mt[presentation[vehicle_id][j]][0])
        result.append(tmpstr)        
        tmpstr = 'Quãng đường của lộ trình: {}km'.format(route_distance)
        result.append(tmpstr)
        tmpstr = 'Khối lượng vận chuyển của lộ trình: {}'.format(nowload)
        result.append(tmpstr)
        tmpstr = "                                     "
        result.append(tmpstr)
    wd_names_now = data['wd_names_now']
    dt = modify_data(data1,False)
    for i in range(len(dt)):
        if dt[i] != 0:
            for v in range(dt[i]):
                vehicle_id+=1
                tmpstr = 'Lộ trình {0} - xe {1} KGS'.format(vehicle_id,data['list_cap'][0])
                result.append(tmpstr)
                tmpstr = 'Từ {0} đến {1}: Khối lượng: {2} KGS | Quãng đường: {3} km'.format(0,i,data1['list_cap'][0],data1['distance_matrix'][0][i])
                result.append(tmpstr)
                tmpstr = 'Từ {0} đến {1}: Khối lượng: {2} KGS | Quãng đường: {3} km'.format(i,0,0,data1['distance_matrix'][i][0])
                # tmpstr = 'Từ {0} Vận chuyển({1} KGS) ->  {2} Vận chuyển({3} KGS) -> {4} Vận chuyển({5} KGS) '.format(0,0,i,data1['list_cap'][0],0,0)
                result.append(tmpstr)
                rt = int(data1['distance_matrix'][0][i] +   data1['distance_matrix'][i][0])
                tmpstr = 'Quãng đường của lộ trình: {}km'.format(rt)
                result.append(tmpstr)
                tmpstr = 'Khối lượng vận chuyển của lộ trình: {}'.format(data1['list_cap'][0])
                result.append(tmpstr)
                total_distance += rt
                tmpstr = '                              '
                result.append(tmpstr)
                total_load += data1['list_cap'][0]
                # write result to result.xlsx
                stt_cell=sheet_obj.cell(row=index_row,column=1)
                stt_cell.value = stt
                month_cell = sheet_obj.cell(row=index_row,column=2)
                x = datetime.datetime.now()
                month_cell.value = x.strftime("%b") + "-" + x.strftime("%y")
                date_cell = sheet_obj.cell(row=index_row,column=3)
                date_cell.value = x.strftime("%d")
                pickUp_cell = sheet_obj.cell(row=index_row,column=4)
                pickUp_cell.value = data['wd_names_now'][0]
                deli_cell = sheet_obj.cell(row=index_row,column=5)
                deli_cell.value = wd_names_now  [i]
                weight_cell = sheet_obj.cell(row=index_row,column=9)
                weight_cell.value = data1['list_cap'][0]
                route_cell = sheet_obj.cell(row = index_row, column = 6 )
                route_cell.value = "Route " + str(vehicle_id+1)
                orderRoute_cell = sheet_obj.cell(row = index_row, column = 7)
                orderRoute_cell.value = 1
                capactiy_cell = sheet_obj.cell(row = index_row, column = 10)
                capactiy_cell.value  = data1['list_cap'][0]
                distance_cell = sheet_obj.cell(row = index_row, column = 8)
                distance_cell.value = rt
                index_row += 1
                stt += 1
                #
    for index_cell in range(1,10):
        space_cell = sheet_obj.cell(row = index_row, column = index_cell)   
        if (index_cell != 8  and index_cell != 1):
            space_cell.value = "#" 
        elif index_cell == 1:
            space_cell.value = "Tổng"
        else:
            space_cell.value = total_distance
    size_cap = len(data1['list_cap'])
    tke_xe = []
    for i in range(size_cap):
        tke_xe.append(0)
    for i in new_cap:
        for j in range(size_cap):
            if i == data1['list_cap'][j]:
                tke_xe[j] += 1
                break
    for i in dt:
        if i != 0:
            tke_xe[0] += i
    
    tmpstr = "Tổng số lượng xe:"
    result.append(tmpstr)
    for i in range(size_cap):
        if tke_xe[i] != 0:
            tmpstr = '   -  {0} xe {1} KGS'.format(tke_xe[i],data1['list_cap'][i])
            result.append(tmpstr)
    tmpstr = '            '
    result.append(tmpstr)
    tmp_str1 = 'Tổng quãng đường di chuyển của tất cả lộ trình: {}km'.format(total_distance)
    result.append(tmp_str1)
    tmp_str2 = 'Tổng khối lượng vận chuyển của tất cả lộ trình: {}'.format(total_load)
    result.append(tmp_str2)
    wb_obj.save("result/result.xlsx")
    return result


def main():
    """Solve the CVRP problem."""
    modify_data(data1,True)
    
    
    bestOne = 10000000000
    for t in range(t_max):
        initPop = population(numChrosome,len(demand) - 1)
        tmpSeque = []
        
        while True:
            tmpSeque = []
            for i in range(num_vehi):
                tmpSeque.append([])
            tmpListLoading = []
            for i in range(num_vehi):
                tmpListLoading.append(0)
            checkCondition = True
            for i in range(1,len(demand)):
                check = False
                for j in range((num_vehi)):
                    if demand[i] + tmpListLoading[j] <=  cap:
                        check = True
                        break
                if check == False:
                    checkCondition = False
                    break
                while True:
                    routeRec = random.randint(0,num_vehi-1)
                    if demand[i] + tmpListLoading[routeRec]  <= cap:
                        tmpSeque[routeRec].append(i)
                        tmpListLoading[routeRec] += demand[i]
                        break
            if checkCondition == True:
                newChromo = chromosome(len(demand)-1,tmpSeque)
                # if newChromo.checkError():
                #     exit()
                initPop.pop.append(newChromo)
            if len(initPop.pop) == numChrosome:
                break
 
        for i in range(generations):
            
            initPop.crossover()

            initPop.mutation()

            initPop.survivorSelection()

            initPop.bestFitness()

            if initPop.bestFit < bestOne:
                
                bestOne = initPop.bestFit
                bestPre = initPop.bestChro.presentation
                # print(bestPre)
                # print(initPop.bestChro.checkError())

    rs = print_solution(data1,bestPre)
    return rs

class Ui_MainWindow(object):
    
    def setupUi(self, MainWindow,widget_names):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(770, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(380, 50, 75, 23))
        self.pushButton.setObjectName("pushButton")
        self.pushButton3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton3.setGeometry(QtCore.QRect(480, 50, 160, 23))
        self.pushButton3.setObjectName("pushButton3")
        self.pushButton4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton4.setGeometry(QtCore.QRect(610, 120, 80, 23))
        self.pushButton4.setObjectName("pushButton3")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(30, 50, 251, 23))
        self.comboBox.setObjectName("comboBox")
        for i in widget_names:
            self.comboBox.addItem("")
        #
        self.comboBox1 = QtWidgets.QComboBox(self.centralwidget)

        self.comboBox1.setGeometry(QtCore.QRect(400, 120, 200, 23))
        self.comboBox1.setObjectName("comboBox1")
        for i in widget_names:
            self.comboBox1.addItem("")
        #
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(400, 100, 47, 13))
        #
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(30, 30, 47, 13))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(290, 30, 100, 13))
        self.label_2.setObjectName("label_2")
        self.spinBox = QtWidgets.QSpinBox(self.centralwidget)
        self.spinBox.setGeometry(QtCore.QRect(290, 50, 80, 22))
        self.spinBox.setObjectName("spinBox")
        self.spinBox.setRange(0,100000)
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(30, 90, 101, 16))
        self.label_3.setObjectName("label_3")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(90, 120, 181, 20))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_2.textChanged.connect(self.update_display)
        #search bar
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(30, 120, 61, 16))
        self.label_4.setObjectName("label_4")
        self.pushButton2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton2.setGeometry(290,120,80,20)
        # #scroll area
        # self.scroll = QtWidgets.QScrollArea(self.centralwidget)
        # self.scroll.setGeometry(QtCore.QRect(30,150,700,430))
        # self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        # self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        # self.scroll.setWidgetResizable(True)  

        #
        self.controls = QWidget(self.centralwidget) #Controls container Widget
        self.controls.setGeometry(QtCore.QRect(30,150,400,400))
        self.controlsLayOut  = QVBoxLayout() # Controls container Layout
        #List of names, widgets, are stored in a dictionary by these keys
        
        self.widgets = []

        #Iterate the names, creating a new OnOffWidget for
        # each one, adding it to the layout and
        # storing a reference in the self.widgets list
        for name in widget_names:
            item = OnOffWidget(name)
            self.controlsLayOut.addWidget(item)
            self.widgets.append(item)
        spacer = QSpacerItem(1, 1, QSizePolicy.Minimum, QSizePolicy.Expanding)
        self.controlsLayOut.addItem(spacer)
        self.controls.setLayout(self.controlsLayOut)

        #Scroll Area Properties
        self.scroll = QScrollArea(self.centralwidget)
        self.scroll.setGeometry(QtCore.QRect(30,150,700,430))
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll.setWidgetResizable(True)        
        self.scroll.setWidget(self.controls)
        
        
        #
        # #Search bar
        # self.searchbar = QLineEdit()
        # self.searchbar.textChanged.connect(self.update_display)

        # # Adding the completer
        # self.completer = QCompleter(widget_names)
        # self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        # self.searchbar.setCompleter(self.completer)
        

        #Add the items to  VBoxLayout (applied to container widget)
        # which encompasses the whole window
        
        #
        # self.spinBox.setKeyboardTracking(False)
        # self.spinBox.editingFinished.connect(self.get_text_in_comboBox)
        self.pushButton.clicked.connect(self.get_text_in_comboBox)
        self.pushButton2.clicked.connect(self.run)
        self.pushButton3.clicked.connect(self.addData)
        self.pushButton4.clicked.connect(self.changeDepot)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 770, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow,widget_names)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow,widget_names):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Màn hình chính"))
        self.pushButton.setText(_translate("MainWindow", "Lựa chọn"))
        self.pushButton3.setText(_translate("MainWindow", "Thêm điểm khách hàng mới"))
        self.pushButton4.setText(_translate("MainWindow", "Đổi kho"))

        for i in range(len(widget_names)):
            self.comboBox.setItemText(i, _translate("MainWindow", widget_names[i]))
        for i in range(len(widget_names)):
            self.comboBox1.setItemText(i, _translate("MainWindow", widget_names[i]))
        self.label.setText(_translate("MainWindow", "Tên "))
        self.label_2.setText(_translate("MainWindow", "Khối lượng giao hàng (KGS):"))
        self.label_3.setText(_translate("MainWindow", "Danh sách yêu cầu:"))
        self.label_4.setText(_translate("MainWindow", "Tìm kiếm:"))
        self.label_5.setText(_translate("MainWindow", "Tên kho:"))

        self.pushButton2.setText(_translate("MainWindow", "Chạy"))
    def changeDepot(self):
        # 
        old_depot = 0
        text = str(self.comboBox1.currentText())
        count = 0
        for i in self.widgets:
            if text == i.name:
                new_depot = count
                break
            count += 1
        
        print(text)
        # open file excel
        wb_obj1 = openpyxl.load_workbook("data/coordinate.xlsx")
        sheet_obj1 = wb_obj1.active
        max_column1=sheet_obj1.max_column
        max_row1=sheet_obj1.max_row
        #
        wb_obj2 = openpyxl.load_workbook("data/data.xlsx")
        sheet_obj2 = wb_obj2["Sheet1"]
        max_column2=sheet_obj2.max_column
        max_row2=sheet_obj2.max_row
        # change row col in distance matrix 
        size = sheet_obj2.max_column-2
        extra_index = 2
        # đổi vị trí những chỗ không phải giao nhau
        for i in range(size):
            if (i != old_depot and i != new_depot):
                tmp = sheet_obj2.cell(row = old_depot+extra_index+1,column=1+i+extra_index).value
                sheet_obj2.cell(row = old_depot+extra_index+1,column=1+i+extra_index).value = sheet_obj2.cell(row = new_depot+extra_index+1,column=1+i+extra_index).value
                sheet_obj2.cell(row = new_depot+extra_index+1,column=1+i+extra_index).value = tmp
                tmp = sheet_obj2.cell(row = i+extra_index+1,column=1+old_depot+extra_index).value
                sheet_obj2.cell(row = i+extra_index+1,column=1+old_depot+extra_index).value =sheet_obj2.cell(row = i+extra_index+1,column=1+new_depot+extra_index).value
                sheet_obj2.cell(row = i+extra_index+1,column=1+new_depot+extra_index).value = tmp
        # đổi vị trí những chỗ giao nhau
        tmp = sheet_obj2.cell(row = old_depot+extra_index+1,column=1+new_depot+extra_index).value
        sheet_obj2.cell(row = old_depot+extra_index+1,column=1+new_depot+extra_index).value = sheet_obj2.cell(row = new_depot+extra_index+1,column=1+old_depot+extra_index).value
        sheet_obj2.cell(row = new_depot+extra_index+1,column=1+old_depot+extra_index).value = tmp
        # đổi tên
        tmp = sheet_obj2.cell(row = 1, column= old_depot+extra_index+1).value
        sheet_obj2.cell(row = 1, column= old_depot+extra_index+1).value = sheet_obj2.cell(row = 1, column= new_depot+extra_index+1).value
        sheet_obj2.cell(row = 1, column= new_depot+extra_index+1).value = tmp
        tmp = sheet_obj2.cell(row = old_depot+extra_index+1, column = 1).value
        sheet_obj2.cell(row = old_depot+extra_index+1, column = 1).value = sheet_obj2.cell(row = new_depot+extra_index+1, column = 1).value
        sheet_obj2.cell(row = new_depot+extra_index+1, column = 1).value  = tmp
        tmp = sheet_obj2.cell(row = 2, column= old_depot+extra_index+1).value
        sheet_obj2.cell(row = 2, column= old_depot+extra_index+1).value = sheet_obj2.cell(row = 2, column= new_depot+extra_index+1).value
        sheet_obj2.cell(row = 2, column= new_depot+extra_index+1).value = tmp
        tmp = sheet_obj2.cell(row = old_depot+extra_index+1, column = 2).value
        sheet_obj2.cell(row = old_depot+extra_index+1, column = 2).value = sheet_obj2.cell(row = new_depot+extra_index+1, column = 2).value
        sheet_obj2.cell(row = new_depot+extra_index+1, column = 2).value  = tmp
        wb_obj2.save("data/data.xlsx")
        # change row in coordinate
        for i in range(3):
            tmp = sheet_obj1.cell(row = old_depot+1,column = i+1).value
            sheet_obj1.cell(row = old_depot+1,column = i+1).value = sheet_obj1.cell(row = new_depot+1,column = i+1).value
            sheet_obj1.cell(row = new_depot+1,column = i+1).value = tmp
        wb_obj1.save("data/coordinate.xlsx")
        global or_data
        or_data = create_data_model()
        countName = 0
        tmpName = data1['wd_names'][old_depot]
        data1['wd_names'][old_depot] = data1['wd_names'][new_depot]
        data1['wd_names'][new_depot] = tmpName
        data1['demands'] = copy.deepcopy(or_data['demands'])
        data1['distance_matrix'] = copy.deepcopy(or_data['distance_matrix'])
        ui.setupUi(MainWindow,data1['wd_names'])
    def addData(self):
        self.w2 = AnotherWindow1()
        self.w2.show()
    def run(self):
        self.w = AnotherWindow()
        self.w.show()
        data1['demands'] = copy.deepcopy(or_data['demands'])
        data1['distance_matrix'] = copy.deepcopy(or_data['distance_matrix'])
        self.update_element()
    def update_element(self):
        for i in range(len(self.widgets)):
            if self.widgets[i].is_on:
                data1['demands'][i] = int(self.widgets[i].lbl1.text())

    def update_display(self, text):
        for widget in self.widgets:
            if text.lower() in widget.name.lower():
                widget.show()
            else:
                widget.hide()
    def get_text_in_comboBox(self):
        text = str(self.comboBox.currentText())
        num = self.spinBox.value()
        count = 0
        for i in self.widgets:
            if text == i.name:
                i.is_on = True
                i.update_button_state()
                i.show()
                data1['demands'][count] = num
                i.update_num(num)
                break   
            count += 1


class OnOffWidget(QWidget):
    def __init__(self,name):
        super(OnOffWidget, self).__init__()
        self.name = name # Name of widget used for searching.
        self.is_on = False # Current state (true=ON, false=OFF)

        self.lbl = QLabel(self.name) #the widget label
        self.lbl1 = QLabel('0') #the widget label1
        self.spnbox = QSpinBox()
        self.spnbox.setRange(0, 100000)
        
        self.btn_on = QPushButton("Sửa") #The ON button
        self.btn_off = QPushButton("Xoá") #the off button
        self.btn_on.clicked.connect(self.update_num_by_spnb) #connect to function On
        self.btn_off.clicked.connect(self.off) #connect to function Off
        self.hbox = QHBoxLayout()  #A horizontal layout  to encapsulate the above
        self.hbox1 = QHBoxLayout() 
        self.hbox2 = QHBoxLayout() 
        self.hbox2.addWidget(self.spnbox) #Add the spin box to the layout
        self.hbox1.addWidget(self.lbl) #Add the label to the layout 
        self.hbox1.addWidget(self.lbl1) #Add the label to the layout 

        self.hbox2.addWidget(self.btn_on)  # Add the ON button to the layout
        self.hbox2.addWidget(self.btn_off)  # Add the OFF button to the layout
        self.hbox.addLayout(self.hbox1)
        self.hbox.addLayout(self.hbox2)
        
        self.setLayout(self.hbox)
        self.update_button_state()
    def update_num_by_spnb(self):
        num = self.spnbox.value()
        for i in range(len(data1['wd_names'])):
            if data1['wd_names'][i] == self.name:
                data1['demands'][i] = num
                break
        self.lbl1.setText(str(num))
    def update_num(self,num):
        self.lbl1.setText(str(num))
    
    def off(self):
        self.is_on = False
        for i in range(len(data1['wd_names'])):
            if data1['wd_names'][i] == self.name:
                data1['demands'][i] = 0
                break
        self.update_button_state()

    def on(self):
        self.is_on = True
        self.update_button_state()
    def update_button_state(self):
        #Update the  appearance of the control buttons(on/off)
        #depending the current state
        if self.is_on == True:
            self.btn_on.setStyleSheet('background-color : #4CA5F0; color : #fff;')
            self.btn_off.setStyleSheet('background-color : none; color: none;')
            self.show()
        else:
            self.btn_off.setStyleSheet('background-color : #D32F2F; color : #fff;')
            self.btn_on.setStyleSheet('background-color : none; color: none;')
            self.hide()
    def show(self):
        # show this widget, and all child widgets
        if self.is_on:
            for w in [self, self.lbl, self.btn_on, self.btn_off, self.spnbox]:
                w.setVisible(True)
    def hide(self):
        # hide this widget, and all childwidgets
        for w in [self, self.lbl, self.btn_on, self.btn_off, self.spnbox]:
            w.setVisible(False)
class AnotherWindow(QMainWindow):
    """
    This "window" is a QWidget. If it has no parent, it
    will appear as a free-floating window as we want.
    """
    def __init__(self, *args, **kwargs):
        super().__init__()

        self.controls = QWidget() #Controls container Widget
        self.controlsLayOut  = QVBoxLayout() # Controls container Layout
        #List of names, widgets, are stored in a dictionary by these keys
        rs = main()
        widget_names = rs
        self.widgets = []
        
        #Iterate the names, creating a new OnOffWidget for
        # each one, adding it to the layout and
        # storing a reference in the self.widgets list
        for name in widget_names:
            item = QLabel(name)
            self.controlsLayOut.addWidget(item)
            self.widgets.append(item)
        spacer = QSpacerItem(1, 1, QSizePolicy.Minimum, QSizePolicy.Expanding)
        self.controlsLayOut.addItem(spacer)
        self.controls.setLayout(self.controlsLayOut)

        #Scroll Area Properties
        self.scroll = QScrollArea()
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll.setWidgetResizable(True)        
        self.scroll.setWidget(self.controls)

        #Add the items to  VBoxLayout (applied to container widget)
        # which encompasses the whole window

        container = QWidget()
        containerLayOut = QVBoxLayout()
        containerLayOut.addWidget(self.scroll)

        container.setLayout(containerLayOut)
        self.setCentralWidget(container)
        
        self.setGeometry(660, 100, 500, 400)
        self.setWindowTitle('Control Panel')
class AnotherWindow1(QMainWindow):
    """
    This "window" is a QWidget. If it has no parent, it
    will appear as a free-floating window as we want.
    """
    def __init__(self, *args, **kwargs):
        super().__init__()
        self.label1 = QLabel("Điểm dữ liệu mới:")
        self.label7 = QLabel("Tên:")
        self.lineEdit5 = QLineEdit()
        self.label2 = QLabel("Kinh độ và Vĩ độ:")
        self.lineEdit1 = QLineEdit()

        self.pushButton1 = QPushButton("Thêm")


        #Add the items to  VBoxLayout (applied to container widget)
        # which encompasses the whole window

        container = QWidget()
        containerLayOut = QVBoxLayout()
        containerLayOut.addWidget(self.label1)
        containerLayOut.addWidget(self.label7)
        containerLayOut.addWidget(self.lineEdit5)
        containerLayOut.addWidget(self.label2)
        containerLayOut.addWidget(self.lineEdit1)

        containerLayOut.addWidget(self.pushButton1)

        
        self.pushButton1.clicked.connect(self.addOnePoint)
        container.setLayout(containerLayOut)
        self.setCentralWidget(container)
        
        self.setGeometry(400, 100, 250, 200)
        self.setWindowTitle('Thêm Dữ liệu')    
    def addOnePoint(self):
        # get input
        nameOfPoint = self.lineEdit5.text()
        laAndLo = self.lineEdit1.text()
        laAndLo = laAndLo.split(',')
        latitude = float(laAndLo[0])
        longitude = float(laAndLo[1])
        #
        data1['wd_names'].append(nameOfPoint)
        ui.setupUi(MainWindow,data1['wd_names'])
        MainWindow.show()
        #
        wb_obj1 = openpyxl.load_workbook("data/coordinate.xlsx")
        sheet_obj1 = wb_obj1.active
        max_column1=sheet_obj1.max_column
        max_row1=sheet_obj1.max_row
        #
        wb_obj2 = openpyxl.load_workbook("data/data.xlsx")
        sheet_obj2 = wb_obj2["Sheet1"]
        max_column2=sheet_obj2.max_column
        max_row2=sheet_obj2.max_row
        #
        sheet_obj2.cell(row= max_row2+1, column = 1).value = nameOfPoint
        sheet_obj2.cell(row= 1, column =max_column2+ 1).value = nameOfPoint
        sheet_obj1.cell(row = max_row1+1 , column = 1).value = nameOfPoint
        sheet_obj1.cell(row = max_row1+1 , column = 2).value = latitude
        sheet_obj1.cell(row = max_row1+1 , column = 3).value = longitude
        # caculate distance by bing map api
        if max_row1 != 0:
            for i in range(1,max_row1+1):
                latitude1 = sheet_obj1.cell(row=i,column=2).value
                longitude1 = sheet_obj1.cell(row=i,column=3).value
                #
                routeUrl = "http://dev.virtualearth.net/REST/V1/Routes/Driving?wp.0=" + str(latitude) + "," + str(longitude) + "&wp.1=" + str(latitude1) + "," + str(longitude1) + "&key=" + bingMapsKey
                request = urllib.request.Request(routeUrl)
                response = urllib.request.urlopen(request)
                r = response.read().decode(encoding="utf-8")
                result = json.loads(r)
                distance1 = ( result["resourceSets"][0]["resources"][0]["routeLegs"][0]["travelDistance"])
                #
                routeUrl = "http://dev.virtualearth.net/REST/V1/Routes/Driving?wp.0=" + str(latitude1) + "," + str(longitude1) + "&wp.1=" + str(latitude) + "," + str(longitude) + "&key=" + bingMapsKey
                request = urllib.request.Request(routeUrl)
                response = urllib.request.urlopen(request)
                r = response.read().decode(encoding="utf-8")
                result = json.loads(r)
                distance2 = ( result["resourceSets"][0]["resources"][0]["routeLegs"][0]["travelDistance"])
                #
                sheet_obj2.cell(row=max_row2+1 , column = i+2).value = distance1
                sheet_obj2.cell(row=i+2 , column = max_column2+ 1).value = distance2
        sheet_obj2.cell(row=max_row2+1 , column =  max_column2+ 1).value = 0
        wb_obj1.save("data/coordinate.xlsx")
        wb_obj2.save("data/data.xlsx")
        global or_data
        or_data = create_data_model()
        data1['demands'] = copy.deepcopy(or_data['demands'])
        data1['distance_matrix'] = copy.deepcopy(or_data['distance_matrix'])
if __name__ == '__main__':

    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow,data1['wd_names'])
    MainWindow.show()

    sys.exit(app.exec_())
